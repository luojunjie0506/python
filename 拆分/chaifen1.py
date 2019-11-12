from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Border,Side,Alignment

545
#写入方法
def cx(a,qq):
    # 新建一个excel文件
    wb = Workbook()
    # 新建一个sheet
    mySheet = wb.create_sheet(index=0, title=a)

    #存储相同值的行
    rank_list = [0]
    for i in range(1, row+1):
        if sheet.cell(row=i,column=num).value == a:
            rank_list.append(i)
    print(rank_list)

    #不显示网格
    mySheet.sheet_view.showGridLines = False

    #设置列宽
    for i in range(1, column+1):
        yy = chr(i+64)
        o =lk[i-1]
        zm1 = yy + str(1)
        zm2 = yy + str(column)
        mySheet.column_dimensions[yy].width = o
        mySheet.merge_cells(zm1+':'+zm2)


    # 写入sheet中
    for hang in range(len(rank_list)): #循环每类的长度
        for lie in range(1,column+1):  #循环列
            if  hang == 0:
                bb = sheet.cell(row=1, column=1).value
                mySheet.cell(row=hang+1, column=lie, value=bb)
            elif hang == 1:
                bb = sheet.cell(row=hang+1, column=lie).value
                mySheet.cell(row=hang+1, column=lie, value=bb)
            else:
                bb = sheet.cell(row=rank_list[hang],column=lie+1).value  #获取原表某行某列的值
                mySheet.cell(row=hang+1, column=1, value=hang)  #新建表每行第一列的序号递增
                mySheet.cell(row=hang+1, column=lie+1, value=bb) #在新建表某行某列的写入值

    #设置文字格式
    for hang in range(0,len(rank_list)):
        for lie in range(1,column+1):
            zm = chr(lie+64)
            lh = zm + str(hang+1)
            mySheet[lh].alignment = Alignment(horizontal='center', vertical='center')
            #设置边框
            if hang == 0: #第一行
                if lie == 1: #第一列
                    mySheet[lh].border = Border(top=Side(style='thin', color="000000"),left=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))
                elif lie == column:#最后一列
                    mySheet[lh].border = Border(top=Side(style='thin', color="000000"),right=Side(style='thin', color="000000"),bottom=Side(style='hair', color="000000"))
                else:
                    mySheet[lh].border = Border(top=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))
            elif hang == len(rank_list)-1:#最后一行
                if lie == 1: #第一列
                    mySheet[lh].border = Border(bottom=Side(style='thin', color="000000"),left=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"))
                elif lie == column:#最后一列
                    mySheet[lh].border = Border(bottom=Side(style='thin', color="000000"),right=Side(style='thin', color="000000"))
                else:
                    mySheet[lh].border = Border(bottom=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"))
            else:
                if lie == 1:#第一列
                    mySheet[lh].border = Border(left=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))
                elif lie == column:#最后一列
                    mySheet[lh].border = Border(right=Side(style='thin', color="000000"),bottom=Side(style='hair', color="000000"))
                else:
                    mySheet[lh].border = Border(right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))

    path = 'D:/表格/'+ qq +'/'+ a  + '.xlsx'
    wb.save(path) #保存每类的excel


if __name__ == '__main__':
    cc = input('请输入要按那列拆分：')
    bb = input('请输入第几个表:')
    dd = input('请输入你的文件名(存储在D盘根目录):')
    #cc='执委'
    #dd='社保关爱计划跟进名单-20191107 最终版 - 副本 (2)(4).xlsx'
    #打开Excel
    file = 'D:/'+dd
    data = load_workbook(file)
    #打开Excel中第一个表
    sheet = data.worksheets[int(bb)]
    #获取行
    row=sheet.max_row
    #获取列
    column = sheet.max_column
    #存储列宽
    lk = []
    #存储列的值
    rank_lm = []

    #获取拆分的是第几列
    for s in range(1,column):
        if sheet.cell(row=1, column=s).value == cc:
            num = s
            print(num)
            break;

    #获取列表加入原表列宽
    for i in range(65, 65+column):
        yy = chr(i)
        liekuan =sheet.column_dimensions[yy].width
        lk.append(liekuan)

    #获取拆分值并调用写入方法
    for i in range(2,row+1):
        if sheet.cell(row=i,column=num).value not in rank_lm:
            rank_lm.append(sheet.cell(row=i,column=num).value)
            a = sheet.cell(row=i,column=num).value
            print(a)
            cx(a,cc)


#执委 申报店 赵冬梅骨干
#社保关爱计划跟进名单-20191107 最终版 - 副本 (2)(4).xlsx