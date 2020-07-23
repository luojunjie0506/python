import os

import pandas as pd
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment

#1制表保存
#2排序保存
#3修改样式

def hz1(zb_name):
    global zb_hang
    # 新建一个excel文件
    wb = Workbook()
    # 新建一个sheet
    mySheet = wb.create_sheet(index=0, title=zb_name)
    path = dir_path +'\\'+ zb_name + '.xlsx'

    #写入第一列信息
    cd = len(frist_xx)
    for chang  in range(1,cd+1):
        mySheet.cell(row=1, column=chang, value=frist_xx[chang-1])

    zb_hang = 1
    for a in range(1, rows):
        bb = table.cell(a, num).value
        if bb == zb_name:
            zb_hang=zb_hang+1
            for cc in range(1, column):
                xx = table.cell(a, cc).value
                mySheet.cell(row=zb_hang, column=cc+1, value=xx)
    wb.save(path)  # 保存每类的excel


def paixu():
    file_list = os.listdir(dir_path)  # 获取文件夹中的所有文件名的列表
    # 遍历文件夹中的所有文件名的列表
    for a in file_list:
        file_path = dir_path + '\\' + a  # 拼接完整路径
        data = pd.read_excel(file_path, sheet_name=0)  # 读取文件夹中的表
        df = pd.DataFrame(data)  # 转换格式
        if cc == '所属执委' or cc =='所属咨委':
            df.sort_values(by=['老人总数'], ascending=(False), inplace=True)
        else:
            df.sort_values(by=['申报店', '申报月次'], ascending=(True, False), inplace=True)
        df.to_excel(file_path, encoding='utf-8', sheet_name=a[:-5], index=None)  # 写入sheet中 index无索引
        print(1)



def geshi():
    file_list = os.listdir(dir_path)  # 获取文件夹中的所有文件名的列表
    for a in file_list:
        path = dir_path + '\\'+a
        data1 = load_workbook(path)
        #打开Excel中第一个表
        mySheet1 = data1.worksheets[int(bb)]
        #不显示网格
        mySheet1.sheet_view.showGridLines = False
        # 获取行
        rows1 = mySheet1.max_row

        if cc == '所属执委' or cc =='所属咨委':
            lk = [6,10,14,14,10,12, 12, 14, 14, 25]
        else:
            lk = [6, 10, 10, 14, 8, 12, 15, 25, 10, 10, 10, 13]

        #设置列宽
        for i in range(1, column+1):
            yy = chr(i+64)
            o =lk[i-1]
            mySheet1.column_dimensions[yy].width = o

        # 设置文字格式
        for hang in range(0, rows1):
            if hang< rows1:
                mySheet1.cell(row=hang+1, column=1, value=hang+1)
            for lie in range(1, column + 1):
                zm = chr(lie + 64)
                lh = zm + str(hang + 1)
                mySheet1[lh].alignment = Alignment(horizontal='center', vertical='center')
                mySheet1[lh].border = Border(top=Side(style='thin', color="000000"),
                                                        left=Side(style='thin', color="000000"),
                                                        right=Side(style='thin', color="000000"),
                                                        bottom=Side(style='thin', color="000000"))

        data1.save(path) #保存每类的excel
        print(2)


if __name__ == '__main__':
    #cc = input('请输入要按那列拆分：')
    #bb = input('请输入第几个表:')
    #filename = input('请输入你的文件名(存储在D盘根目录):')
    bb= 0

    # 所属执委 所属咨委 02 未参与店铺名单（正常店）.xlsx
    # 执委，咨委，03 未参与老人名单.xlsx
    list = ['所属执委','所属咨委', '执委','咨委']
    for zb_x in list:
        cc = zb_x
        if cc == '所属执委' or cc =='所属咨委':
            filename = '02 未参与店铺名单（正常店）.xlsx'
        else:
            filename = '03 未参与老人名单.xlsx'
        file_path = 'D:/'+filename
        data = xlrd.open_workbook(file_path)
        table = data.sheet_by_index(bb)
        #获取行
        rows = table.nrows
        #获取列
        column = table.ncols
        zb_list = [] #存放拆分值名字
        frist_xx = [] #存放第一列信息

        #创建文件夹存放各店信息
        dir_path = 'D:\\表格\\' +  filename[3:-5] + '-'+cc
        os.mkdir(dir_path)

        #查询拆分值在表中的第几列
        for s in range(0,column):
            zd = table.cell(0, s).value
            frist_xx.append(zd)
            if zd == cc:
                num = s
                print(num)

        #获取拆分值并调用写入方法
        for a in range(1,rows):
            zb_name = table.cell(a,num).value
            if zb_name not in zb_list:
                zb_list.append(zb_name)
                hz1(zb_name)

        paixu()
        geshi()


