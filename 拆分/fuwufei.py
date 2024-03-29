# coding=gbk
import datetime
import os
import pandas as pd
import xlrd
from openpyxl import *
import numpy as np
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def getmonth():
    # 获取当前月份-1
    year = datetime.datetime.now().year
    month = datetime.datetime.now().month
    if month ==1 :
        ym = str(year-1) + str(12)
    elif month < 10:
        ym = str(year) + '0' + str(month - 1)
    else:
        ym = str(year) + str(month - 1)
    return ym

# 整合文件夹中的表到一个excel中
def zhenghe():
    global sheets_list
    path = 'D:\\fuwufei3\\c\\' # 文件夹路径
    save_path = 'D:\\fuwufei3\\xx.xlsx'  # 合并后excel的存放路径
    file_list = os.listdir(path)  # 获取文件夹中的所有文件名的列表
    writer = pd.ExcelWriter(save_path)  # 使用ExcelWriter函数，可以写入数据时不会覆盖sheet
    # 遍历文件夹中的所有文件名的列表
    for a in file_list:
        file_path = path + a  # 拼接完整路径
        data = pd.read_excel(file_path, sheet_name=0)  # 读取文件夹中的表
        df = pd.DataFrame(data)  # 转换格式

        # 循环每个表中表头的第一个字段，并用来排序
        for btname in data:
            #判断表是否是服务费清单原始数据表，是的话就要表中第一个字段和顺序字段排列
            if '原始数据' in a:
                df.sort_values(by=[btname, '身份'], ascending=True, inplace=True)  # 使用第一个字段来升序。ascending排序，inplace替代
            else:
                df.sort_values(by=btname, ascending=True, inplace=True)  # 使用第一个字段来升序。ascending排序，inplace替代
            df.to_excel(writer, encoding='utf-8', sheet_name=a[:-5], index=None)  # 写入sheet中 index无索引
            writer.save()
            writer.close()
            sheets_list.append(a[:-5])
            break
    print(sheets_list)
    return save_path

def xhs(dh):
    global b0, b1, b2, b3, b4,b5
    global sheet0, sheet1, sheet2, sheet3, sheet4,sheet5
    global row_list, col_list
    list1 = []  # 存储匹配服务费清单原始数据
    list2 = []  # 存储匹配补扣款备注,小微店补,门店信息表
    list3 = []  # 存储匹配活动明细表
    list4 = []  # 存储匹配业绩明细表
    # 找到每个表的行数，如有存在行数就追加到list1,如不存在，在list1中加0后到后一个表匹配的行数
    # b参数是记录每次查询到的位置
    for a0 in range(b0, row_list[0]):
        mh0 = []
        # b的位置跟长度相等时，说明表中数据一匹配完，直接在list中加0
        if a0 == row_list[0]:
            break
        else:
            cell_0 = sheet0.cell(a0, 0).value
            if cell_0 != dh:
                break
            else:
                for index in range(0,col_list[0]):
                    value_0 = sheet0.cell(a0, index).value
                    mh0.append(value_0)
                list1.append(mh0)
                b0 = a0 + 1

    # for循环中如果range中的两个参数都一样时,那这个循环就不执行,为了让循环继续执行所以+1
    for a1 in range(b1, row_list[1]+1):
        if a1 == row_list[1]:
            list2.append('')
            break
        else:
            cell_1 = sheet1.cell(a1, 0).value
            if cell_1 == dh:
                for index in range(0,col_list[1]):
                    value_1 = sheet1.cell(a1, index).value
                    list2.append(value_1)
                b1 = a1 + 1
                break

    for a2 in range(b2, row_list[2]+1):
        if a2 == row_list[2]:
            list2.append('')
            break
        else:
            cell_2 = sheet2.cell(a2, 0).value
            if cell_2 == dh:
                value_2 = sheet2.cell(a2, 2).value
                list2.append(value_2)
                b2 = a2 + 1
                break

    for a3 in range(b3, row_list[3]+1):
        if a3 == row_list[3]:
            list2.append('')
            break
        else:
            cell_3 = sheet3.cell(a3, 0).value
            if cell_3 == dh:
                value_3 = sheet3.cell(a3, 3).value
                list2.append(value_3)
                b3 = a3 + 1
            else:
                break

    for a4 in range(b4, row_list[4]):
        mh1 = []
        if a4 == row_list[4]:
            break
        else:
            cell_4 = sheet4.cell(a4, 0).value
            if cell_4 != dh:
                break
            else:
                for index in range(0,col_list[4]):
                    value_4 = sheet4.cell(a4, index).value
                    mh1.append(value_4)
                list3.append(mh1)
                b4 = a4 + 1

    for a5 in range(b5, row_list[5]):
        mh2 = []
        if a5 == row_list[5]:
            break
        else:
            cell_5 = sheet5.cell(a5, 0).value
            if cell_5 == dh:
                for index in range(0,col_list[5]):
                    value_5 = sheet5.cell(a5, index).value
                    mh2.append(value_5)
                list4.append(mh2)
                b5 = a5 + 1

    #数据处理
    list1  = dataCl1(list1,list2)
    list3 = dataCl3(list3)
    # 传入xr方法
    xr(list1,list2,list3,list4)

# list1中数据的处理
def dataCl1(list1,list2):
    def take2(elem):
        return elem[1]
    list1.sort(key=take2)
    sum = 0
    for yiwei in range(0, len(list1)):
        # 判断list1的第一行是否店长
        if list1[yiwei][1] == '店长' and yiwei == 0:
            if list2[5] != '':
                # list2中小薇店补不为空,加一行小薇,一行合计
                sss = []
                sss.append(list2[0])
                sss.append(list2[1])
                sss.append(list2[3])
                sss.append(list2[4])
                sss.append('小微店补')
                sss.append(list2[5])
                list1.insert(1, sss)
                sss1 = ['电子积分个人账户入账合计']
                nuM = float(list1[0][14]) + float(list2[5])
                sss1.append(nuM)
                sss1.append(list1[0][18])
                list1.insert(2, sss1)
            else:
                # list2中小薇店补为空,加一行合计
                sss1 = ['电子积分个人账户入账合计']
                sss1.append(list1[0][14])
                sss1.append(list1[0][18])
                list1.insert(1, sss1)
        elif list1[yiwei][1] != '店长' and yiwei == 0:
            if list2[5] != '':
                sss = []
                sss.append(list2[0])
                sss.append(list2[1])
                sss.append(list2[3])
                sss.append(list2[4])
                list1.insert(0, sss)
                sss1 = ['电子积分个人账户入账合计']
                sss1.append(list2[5])
                list1.insert(1, sss1)

    for num in range(1, len(list1)):
        zhi1 = list1[num - 1][1]
        zhi2 = list1[num][1]
        if zhi1 == '门店代发＜200元':
            if zhi1 != zhi2:
                sum += float(list1[num - 1][17])
                sss1 = ['电子积分代发账户入账合计']
                sss1.append(sum)
                sss1.append('代发账户')
                list1.insert(num, sss1)
                break
            elif zhi1 == zhi2:
                if num == len(list1) - 1:
                    sum += float(list1[num - 1][17])
                    sum += float(list1[num][17])
                    sss1 = ['电子积分代发账户入账合计']
                    sss1.append(sum)
                    sss1.append('代发账户')
                    list1.insert(num + 1, sss1)
                    break
                else:
                    sum += float(list1[num - 1][17])

    for i in range(0,len(list1)):
        list1[i].insert(0, i+1)
    value = ''
    for x in range(0, len(list1)):
        if list1[x][2] == '门店经销商≥200元':
            list1[x][2] = '门店经销商'
        # print(list1[x][2])
        if list1[x][2] != value:
            value = list1[x][2]
        else:
            list1[x][2] = ''
    return list1

# 设置超链接
def link_1(cell, link, display=None):
    cell.hyperlink = link
    cell.font = Font(size=9, bold=True, name='微软雅黑',color='336600')
    if display is not None:
        cell.value = display

# list3数据处理
def dataCl3(list3):
    if len(list3) > 0:
        list4 = np.array(list3)
        idex = np.lexsort([list4[:, 1], list4[:, 5]])
        list3 = list4[idex, :]
        list3 = list3.tolist()
        sum = 0.0 #存每个人的合计
        ss = [] #存第几行插入合计
        for i in range(0, len(list3)):
            if i == len(list4) - 1:
                sum += float(list3[i][6])
                ss.append([i + 1, sum])
                break
            elif list3[i][5] != list3[i + 1][5]:
                sum += float(list3[i][6])
                ss.append([i + 1, sum])
                sum = 0.0
            else:
                sum += float(list3[i][6])
        for y in range(0, len(ss)):
            cc = ['合计：']
            cc.append(ss[y][1])
            # 因为循环了几次就插入了几行,行数变化所以加上y
            list3.insert(ss[y][0] + y, cc)
        value = ''
        for x in range(0, len(list3)):
            if list3[x][1] != value:
                value = list3[x][1]
            else:
                list3[x][1] = ''
        return list3
    else:
        return list3


# 新建模板并写入数据
def xr(list1, list2,list3,list4):
    global yue,wcnum,num1,num,v
    xiaowei = [1,3,4,5,6,8,15]
    xw_heji= [1,3,15,19]
    heji = [1,3,18,19]
    heji1 = [1, 3, 15]
    df_row = 0
    gr_row = 0
    wb = load_workbook('D:\\fuwufei3\\清单模板.xlsx')
    ws = wb["清单"]
    ws1 = wb['活动奖项明细']
    ws2 = wb['业绩明细']

    # 业绩明细sheet写入数据
    for cd1 in range(0,len(list4)):
        for cd2 in range(0,len(list4[cd1])):
            if cd2 > 0:
                ws2.cell(row=cd1 + 3, column=2).value = cd1 +1
                ws2.cell(row=cd1 + 3, column=cd2 + 2).value = list4[cd1][cd2]
    # 业绩明细sheet备注
    ws2.cell(row=len(list4)+3,column=2).value = '注：业绩累计实时变动，仅供参考，请以系统为准！'


    #活动奖项明细sheet写入数据
    if len(list3) == 1:
        if len(list3[0]) != 0:
            for i in range(0, len(list3[0])-1):
                ws1.cell(row=3, column=2).value = i + 1
                ws1.cell(row=3, column=i+3).value = list3[0][i+1]
    else:
        for i in range(0, len(list3)):
            if len(list3[i]) >2:
                for c in range(0, len(list3[i])-1):
                    ws1.cell(row=i+3, column=2).value = i + 1
                    ws1.cell(row=i+3, column=c+3).value = list3[i][c+1]
            else:
                ws1.cell(row=i + 3, column=2).value = list3[i][0]
                ws1.cell(row=i + 3, column=8).value = list3[i][1]

    a = '注：服务费发放说明：\n1、本月总收入 = 个人销售奖 + 销售服务奖 + 活动奖 + 领导奖 + 卓越奖（绿色标注：收入构成）\n2、本月总收入=秒结入账+月结入账（紫色标注：发放途径）\n3、活动奖：全拓直接服务奖 + 分享有礼直接服务奖 + 翻倍服务奖（明细详见附件2）\n4、秒结入账：符合秒结规则的分享有礼直接服务奖和对应个人销售奖，已实时入账相应电子积分秒结账户，不再做月结入账\n5、绿色标注代表收入构成；紫色标注代表总收入和发放途径；粉色标注代表本月支出\n6、电子积分账户余额受时时订货、转账等因素影响，请以系统为准，可查阅账户流水'
    n = 6  #清单sheet从哪行开始写
    #清单sheet写入数据
    for yiwei in range(0, len(list1)):
        if len(list1[yiwei]) >10:
            ss = 1#清单sheet从哪列开始写
            for i in range(0,len(list1[yiwei])):
                if i==1:
                    continue
                elif i ==10 and yiwei==0 and list1[0][2] == '店长':
                    link_1(ws['J6'], '#活动奖项明细!A1', list1[yiwei][i])
                    ss += 1
                else:
                    ws.cell(row=n, column=ss).value = list1[yiwei][i]
                    ss += 1
            n +=1
        elif len(list1[yiwei]) == 7:
            ws.cell(row=n, column=2).value = ''
            for xw in range(0,len(xiaowei)):
                ws.cell(row=n, column=xiaowei[xw]).value = list1[yiwei][xw]
            n += 1
        elif len(list1[yiwei]) == 4  :
            ws.cell(row=n, column=2).value = ''
            if list1[yiwei][1] == '电子积分个人账户入账合计':
                paixu = xw_heji
                gr_row = yiwei
            else:
                paixu = heji
                df_row = yiwei
            for hj in range(0, len(paixu)):
                ws.cell(row=n, column=paixu[hj]).value = list1[yiwei][hj]
            n += 1

        elif  len(list1[yiwei]) == 3 :
            ws.cell(row=n, column=2).value = ''
            for hj1 in range(0,len(heji1)):
                ws.cell(row=n, column=heji1[hj1]).value = list1[yiwei][hj1]
            n += 1


    ws['C3'] = list2[0]
    ws['H3'] = list2[1]
    ws['M3'] = yue
    ws['Q3'] = list2[2]


    #注释文字样式
    cs1 = 'A'+ str(n)
    cs2 ='U' + str(n)
    ws.merge_cells(cs1+':'+cs2) #合并单元格
    ws.row_dimensions[n].height = 130 #设置行高
    ws.cell(row=n, column=1).value = a
    ws.cell(row=n, column=1).font = Font(size=9, bold=True, name='微软雅黑') # 注释文字格式
    ws.cell(row=n, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    #清单sheet格式设置
    font = Font(size=9, bold=False, name='微软雅黑')  # 普通文字格式
    font1 = Font(size=9, bold=True, name='微软雅黑',color='336600') # 收入文字格式
    font2 = Font(size=9, bold=True, name='微软雅黑', color='0000cc')  # 代发收入文字格式
    fill1 = PatternFill(patternType="solid", start_color="eaf1dd")  # 收入底色
    fill2 = PatternFill(patternType="solid", start_color="f2dddc")  # 扣除底色
    fill3 = PatternFill(patternType="solid", start_color="dbe5f1")  # 合计底色
    fill4 = PatternFill(patternType="solid", start_color="e5e0ec")  # 入账底色 浅紫
    fill5 = PatternFill(patternType="solid", start_color="ccc0da")  # 入账底色 深紫

    # 整体字体和居中,颜色填充,合并
    for a1 in range(0,len(list1)):
        ws.row_dimensions[a1+6].height = 24 #设置每行行高
        if len(list1[a1]) == 7:
            for xw in range(0, len(xiaowei)):
                ws.cell(row=a1 + 6, column=xiaowei[xw]).font = font
                ws.cell(row=a1 + 6, column=xiaowei[xw]).alignment = Alignment(horizontal='center', vertical='center')
                cs1 = 'H' + str(a1 + 6)
                cs2 = 'N' + str(a1 + 6)
                cs3 = 'P' + str(a1 + 6)
                cs4 = 'R' + str(a1 + 6)
                ws.merge_cells(cs1 + ':' + cs2)  # 合并单元格
                ws.merge_cells(cs3 + ':' + cs4)  # 合并单元格
        elif len(list1[a1]) == 4 :
            if list1[a1][1] == '电子积分个人账户入账合计':
                paixu = xw_heji
            else:
                paixu = heji
            for hj in range(0, len(paixu)):
                ws.cell(row=a1 + 6, column=paixu[hj]).font = font
                ws.cell(row=a1 + 6, column=paixu[hj]).alignment = Alignment(horizontal='center', vertical='center')
                if list1[a1][1] == '电子积分个人账户入账合计':
                    cs1 = 'C' + str(a1 + 6)
                    cs2 = 'N' + str(a1 + 6)
                    cs3 = 'P' + str(a1 + 6)
                    cs4 = 'R' + str(a1 + 6)
                    ws.merge_cells(cs1 + ':' + cs2)  # 合并单元格
                    ws.merge_cells(cs3 + ':' + cs4)  # 合并单元格
                    for ccc in range(0,a1+1):
                        ws.cell(row=ccc + 6, column=15).fill = fill5
                    ws.cell(row=a1 + 6, column=16).fill = fill3
                else:
                    cs1 = 'C' + str(a1 + 6)
                    cs2 = 'Q' + str(a1 + 6)
                    ws.merge_cells(cs1 + ':' + cs2)  # 合并单元格
                if hj != 0:
                    ws.cell(row=a1 + 6, column=paixu[hj]).fill = fill3
        elif  len(list1[a1]) == 3:
            for hj1 in range(0, len(heji1)):
                ws.cell(row=a1 + 6, column=heji1[hj1]).font = font
                ws.cell(row=a1 + 6, column=heji1[hj1]).alignment = Alignment(horizontal='center', vertical='center')
                cs1 = 'C' + str(a1 + 6)
                cs2 = 'Q' + str(a1 + 6)
                ws.merge_cells(cs1 + ':' + cs2)  # 合并单元格
                if hj1 != 0:
                    ws.cell(row=a1 + 6, column=heji1[hj1]).fill = fill3
        else:
            for i in range(0, len(list1[a1])):
                ws.cell(row=a1 + 6, column=i + 1).font = font
                ws.cell(row=a1+6, column=i+1).alignment = Alignment(horizontal='center', vertical='center')
                if 7<=i<=11:
                    ws.cell(row=a1 + 6, column=i + 1).fill = fill1
                elif 12<=i<=14:
                    ws.cell(row=a1 + 6, column=i + 1).fill = fill4
                elif 15<=i<=17:
                    ws.cell(row=a1 + 6, column=i + 1).fill = fill2

        if a1 > (df_row):
            ws.cell(row=a1 + 6, column=15).fill = fill5

    # 清单表相同的合并
    zhi = ''
    num3 = 6
    num4 = 0
    for a1 in range(0, len(list1)+1):
        if a1 == 0:
            zhi = ws.cell(row=6, column=2).value
        # 合并相同的单元格
        if a1 > 0:
            value = ws.cell(row=a1 + 6, column=2).value
            if value != '':
                zhi1 = value
                num4 = a1 + 6
                if num4 - num3 > 1:
                    nb1 = 'B' + str(num3)
                    nb2 = 'B' + str(num4 - 1)
                    ws.merge_cells(nb1 + ':' + nb2)  # 合并单元格
                    zhi = zhi1
                    num3 = num4

    tc = [13,14,15,18]
    #收入和入账字体颜色
    for a1 in range(0,len(list1)):
        for a2 in range(len(tc)):
            ws.cell(row=a1 + 6, column=tc[a2]).font = font1
    ws.cell(row=df_row+6, column=18).font = font2
    if df_row !=0:
        for nb in range(gr_row,df_row+1):
            ws.cell(row=nb + 6, column=18).fill = fill5

    # 清单边框
    border = Border(top=Side(border_style='thin', color='000000'),bottom=Side(border_style='thin', color='000000'),left=Side(border_style='thin', color='000000'),right=Side(border_style='thin', color='000000'))
    for x in range(0,len(list1)):
        for y in range(0, 21):
            ws.cell(row=x + 6, column=y+1).border = border

     #扣款备注
    for yiwei in range(6, len(list1)+6):
        for uu in range(0, len(list2)):
            if uu > 5 and list2[uu] != '' :
                cardNo = str(list2[uu][0:7])
                cardNo2 = str(list2[uu][0:8])
                cardNo3 = str(list2[uu][0:17])
                cardNo1 = str(ws.cell(row=yiwei, column=3).value)
                if cardNo1 == cardNo or cardNo1 == cardNo2 or cardNo1 == cardNo3:
                    ws.cell(row=yiwei, column=20).value = list2[uu]
                    ws.cell(row=yiwei , column=20).font = Font(size=6, bold=False, name='微软雅黑')
                    ws.cell(row=yiwei , column=20).alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    # 活动奖项明细sheet格式设置
    # 整体字体和居中,颜色填充,合并
    if len(list3) == 1:
        if len(list3[0]) != 0:
            for i in range(0, len(list3[0])-1):
                ws1.cell(row=3, column=2).font = font
                ws1.cell(row=3, column=i+3).alignment = Alignment(horizontal='center', vertical='center')
    else:
        v = 0
        for i in range(0, len(list3)):
            ws1.row_dimensions[i+3].height = 24  # 设置每行行高
            if len(list3[i]) >2:
                for c in range(0, len(list3[i])-1):
                    ws1.cell(row=i + 3, column=2).font = font
                    ws1.cell(row=i + 3, column=c + 3).font = font
                    ws1.cell(row=i + 3, column=2).alignment = Alignment(horizontal='center', vertical='center')
                    ws1.cell(row=i + 3, column=c + 3).alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws1.cell(row=i + 3, column=2).font = font
                ws1.cell(row=i + 3, column=8).font = font
                ws1.cell(row=i + 3, column=2).alignment= Alignment(horizontal='center', vertical='center')
                ws1.cell(row=i + 3, column=8).alignment = Alignment(horizontal='center', vertical='center')
                cs3 = 'B' + str(i + 3)
                cs4 = 'G' + str(i + 3)
                cs5 = 'H' + str(i + 3)
                cs6 = 'I' + str(i + 3)
                ws1.merge_cells(cs3 + ':' + cs4)  # 合并单元格
                ws1.merge_cells(cs5 + ':' + cs6)
                ws1.cell(row=i + 3, column=2).fill = fill3
                ws1.cell(row=i + 3, column=8).fill = fill3
            # 合并相同的单元格
            if i == 0:
                num = 'C3'
            else:
                if list3[i][1] != '':
                    if len(list3[i]) < 5:
                        num1 = 'C' + str(i + 2)
                        num2 = 'C' + str(i + 4)
                        ws1.merge_cells(num + ':' + num1)
                        num = num2
                        v = 1
                    else:
                        if v == 1:
                            v = 0
                            continue
                        else:
                            num1 = 'C' + str(i + 2)
                            num2 = 'C' + str(i + 3)
                            if num != num1:
                                ws1.merge_cells(num + ':' + num1)
                            num = num2

    #活动奖项明细sheet边框
    for x in range(0,len(list3)):
        for y in range(0, 8):
            ws1.cell(row=x + 3, column=y+2).border = border

    # 业绩明细sheet格式
    for x in range(0,len(list4)):
        ws2.row_dimensions[x + 3].height = 20
        for y in range(0,len(list4[cd1])):
            ws2.cell(row=x + 3, column=y + 2).font = font
            ws2.cell(row=x + 3, column=y + 2).alignment = Alignment(horizontal='center', vertical='center')
            ws2.cell(row=x + 3, column=y + 2).border = border
    ws2.cell(row=len(list4) + 3, column=2).font = font

    file_name = dir_path + '\\'+ list2[0] + '.xlsx'
    wb.active = 0  # 设置每次打开excel都是第一个sheet
    wb.save(file_name)
    wb.close()
    wcnum += 1
    print(wcnum)


if __name__ == '__main__':
    b0 = 1
    b1 = 1
    b2 = 1
    b3 = 1
    b4 = 1
    b5 = 1
    wcnum = 0 #完成的数量
    row_list = []  # 存放每个表的行数
    col_list = []  # 存放每个表的列数
    yue = getmonth()
    sheets_list = []  # 存放合并后的sheet名列表
    # sheets_list=['202105 业绩明细', '202105 小微店补', '202105 服务费清单原始数据', '202105 活动明细表', '202105 补扣款备注', '202105 门店信息表']
    zhenghe_path = zhenghe()

    #创建文件夹存放各店信息
    dir_path = 'D:\\fuwufei3\\' + getmonth() + '服务费清单'
    os.mkdir(dir_path)

    data = xlrd.open_workbook("D:\\fuwufei3\\xx.xlsx")
    # 按顺序打开各个sheet表并获取行列
    for num in range(0, len(sheets_list)):
        for sheet_list in sheets_list:
            if ('原始数据' in sheet_list) and num == 0:
                sheet0 = data.sheet_by_name(sheet_list)
                rows0 = sheet0.nrows
                cols0 = sheet0.ncols
                row_list.append(rows0)
                col_list.append(cols0)
                break
            elif ('门店信息' in sheet_list) and num == 1:
                sheet1 = data.sheet_by_name(sheet_list)
                rows1 = sheet1.nrows
                cols1 = sheet1.ncols
                row_list.append(rows1)
                col_list.append(cols1)
                break
            elif ('小微店补' in sheet_list) and num == 2:
                sheet2 = data.sheet_by_name(sheet_list)
                rows2 = sheet2.nrows
                cols2 = sheet2.ncols
                row_list.append(rows2)
                col_list.append(cols2)
                break
            elif ('补扣款备注' in sheet_list) and num == 3:
                sheet3 = data.sheet_by_name(sheet_list)
                rows3 = sheet3.nrows
                cols3 = sheet3.ncols
                row_list.append(rows3)
                col_list.append(cols3)
                break
            elif ('活动明细表' in sheet_list) and num == 4:
                sheet4 = data.sheet_by_name(sheet_list)
                rows4 = sheet4.nrows
                cols4 = sheet4.ncols
                row_list.append(rows4)
                col_list.append(cols4)
                break
            elif ('业绩明细' in sheet_list) and num == 5:
                sheet5 = data.sheet_by_name(sheet_list)
                rows5 = sheet5.nrows
                cols5 = sheet5.ncols
                row_list.append(rows5)
                col_list.append(cols5)
                break

    print(row_list)
    for xhdh in range(1, row_list[0] ):
        dh1 = sheet0.cell(xhdh-1, 0).value
        dh2 = sheet0.cell(xhdh, 0).value # 需要制作清单的店号
        if dh1 != dh2 :
            xhs(dh2)

