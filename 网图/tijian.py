# 整合文件夹中的表到一个excel中
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment, Border, Side


def cx(a,qq):
    # 新建一个excel文件
    wb = Workbook()
    # 新建一个sheet
    mySheet = wb.create_sheet(index=0, title=a)

    #存储相同值的行
    rank_list = [0]
    for i in range(1, row+1):
        if sheet.cell(row=i,column=num).value == a:
            rank_list.append(i)

    print(rank_list)

    #不显示网格
    mySheet.sheet_view.showGridLines = False

    fontStyle = Font(name=u'微软雅黑', size=8)
    fontStyle1 = Font(name=u'微软雅黑', size=5)

    # 写入sheet中
    for hang in range(len(rank_list)): #循环每类的长度
        mySheet.row_dimensions[hang+1].height = 13.5
        for lie in range(2,column+1):  #循环列
            if  hang == 0:
                mySheet.cell(row=1, column=1, value='序号').font = fontStyle
                bb = sheet.cell(row=1, column=lie).value
                mySheet.cell(row=hang+1, column=lie, value=bb).font = fontStyle
            else:
                mySheet.cell(row=hang+1, column=1, value=hang).font = fontStyle
                bb = sheet.cell(row=rank_list[hang],column=lie).value  #获取原表某行某列的值
                mySheet.cell(row=hang+1, column=lie, value=bb).font = fontStyle #在新建表某行某列的写入值

    #设置文字格式
    for hang in range(0,len(rank_list)):
        for lie in range(1,column+1):
            zm = chr(lie+64)
            lh = zm + str(hang+1)
            mySheet[lh].alignment = Alignment(horizontal='center', vertical='center')
            #设置边框
            if hang == 0: #第一行
                if lie == 1: #第一列
                    mySheet[lh].border = Border(top=Side(style='thin', color="000000"),left=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))
                elif lie == column:#最后一列
                    mySheet[lh].border = Border(top=Side(style='thin', color="000000"),right=Side(style='thin', color="000000"),bottom=Side(style='hair', color="000000"))

                else:
                    mySheet[lh].border = Border(top=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))
            elif hang == len(rank_list)-1:#最后一行
                if lie == 1: #第一列
                    mySheet[lh].border = Border(bottom=Side(style='thin', color="000000"),left=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"))
                elif lie == column:#最后一列
                    mySheet[lh].border = Border(bottom=Side(style='thin', color="000000"),right=Side(style='thin', color="000000"))
                    mySheet[lh].font = fontStyle1
                    mySheet[lh].alignment = Alignment(wrapText=True)
                else:
                    mySheet[lh].border = Border(bottom=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"))
            else:
                if lie == 1:#第一列
                    mySheet[lh].border = Border(left=Side(style='thin', color="000000"),right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))
                elif lie == column:#最后一列
                    mySheet[lh].border = Border(right=Side(style='thin', color="000000"),bottom=Side(style='hair', color="000000"))
                    mySheet[lh].font = fontStyle1
                    mySheet[lh].alignment = Alignment(wrapText=True)
                else:
                    mySheet[lh].border = Border(right=Side(style='hair', color="000000"),bottom=Side(style='hair', color="000000"))


    path = dir_path+'/' + str(qq) +'/'+ str(a)  + '.xlsx'
    wb.save(path) #保存每类的excel



if __name__ == '__main__':
    # cc = input('请输入要按那列拆分：')
    dd = '7.03-7.05荣格云健康检测报告名单及门店情况（山西阳泉 孟县）_20210706.xlsx'
    # ['所属执委', '所属咨委', 骨干']
    #创建文件夹存放各店信息
    dir_path ='D:/表格/' + dd[:-5]
    os.mkdir(dir_path)
    #打开Excel
    file = 'D:/'+dd
    data = load_workbook(file)
    #打开Excel中第一个表
    sheet = data.worksheets[0]
    #获取行
    row=sheet.max_row
    #获取列
    column = sheet.max_column


    #存储列宽
    lk = []
    aa_list = [ '开设店', '骨干']
    for i in aa_list:
        dir_path1 = dir_path + '/' + i
        os.mkdir(dir_path1)

    for cc in aa_list:
        # 存储列的值
        rank_lm = []
        #获取拆分的是第几列
        for s in range(1,column):
            if sheet.cell(row=1, column=s).value == cc:
                num = s
                print(num)
                break;


        #获取拆分值并调用写入方法
        for i in range(2,row+1):
            if sheet.cell(row=i,column=num).value not in rank_lm:
                rank_lm.append(sheet.cell(row=i,column=num).value)
                a = sheet.cell(row=i,column=num).value
                print(a)
                cx(a,cc)
